import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ========== ใส่ TOKEN ใหม่หลัง ROTATE ==========
TOKEN = "  รอใส่ TOKEN จากผู้ต้องการเข้าถึงข้อมูล  "
# ================================================

URL = "https://gateway.api.bot.or.th/Stat-ExchangeRate/v2/DAILY_AVG_EXG_RATE/"

headers = {
    "Accept": "*/*",
    "Authorization": TOKEN
}

all_data = []

# กำหนดช่วงเวลาที่ต้องการ
start_date = datetime(2014, 1, 1)
#start_date = datetime(2026, 5, 17)เอาไว้เทส
end_date   = datetime.today()

print("🚀 กำลังดึงข้อมูล THB/USD จาก BOT API...")
print("=" * 45)

# วนลูปทีละ 31 วัน
current = start_date
while current <= end_date:
    period_start = current
    period_end   = min(current + timedelta(days=30), end_date)  # ไม่เกิน 31 วัน

    params = {
        "start_period": period_start.strftime("%Y-%m-%d"),
        "end_period":   period_end.strftime("%Y-%m-%d"),
        "currency":"USD"
    }

    try:
        response = requests.get(URL, headers=headers, params=params, timeout=30)
        data = response.json()

        # ดึงข้อมูลออกมาจาก JSON
        records = data.get("result", {}).get("data", {}).get("data_detail", [])

        for r in records:
            all_data.append({
                "วันที่": r.get("period"),
                "ซื้อเงินโอน (Buying Transfer)": r.get("buying_transfer")
            })

        print(f"  ✅ {params['start_period']} ถึง {params['end_period']}: {len(records)} รายการ")

    except Exception as e:
        print(f"  ❌ {params['start_period']} ถึง {params['end_period']}: {e}")

    current = period_end + timedelta(days=1)  # เลื่อนไปช่วงถัดไป
    #time.sleep(0.5)  # หน่วงนิดนึงไม่ให้ API หนักเกินไป

# ====== รวมข้อมูลและบันทึก Excel ======
print("=" * 45)
df = pd.DataFrame(all_data)
df["วันที่"] = pd.to_datetime(df["วันที่"])
df["ซื้อเงินโอน (Buying Transfer)"] = pd.to_numeric(
    df["ซื้อเงินโอน (Buying Transfer)"], errors="coerce"
)
df = df.sort_values("วันที่").reset_index(drop=True)
df = df.dropna()  # ลบวันที่ไม่มีข้อมูล (วันหยุด)

output = "THB_USD_Historical_12Years.xlsx"
df.to_excel(output, index=False, sheet_name="Exchange Rate")

# เปิดไฟล์มาแปลงเป็น Table
wb = load_workbook(output)
ws = wb["Exchange Rate"]

# กำหนดขนาด Table (ครอบคลุมทุกแถวและคอลัมน์)
last_row = len(df) + 1  # +1 เพราะมี Header
last_col = get_column_letter(len(df.columns))
table_range = f"A1:{last_col}{last_row}"

# สร้าง Table
table = Table(
    displayName="ExchangeRateData",
    ref=table_range
)

# เลือกสไตล์ตาราง (เปลี่ยนเลขท้ายได้ 1-21)
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",  # สีฟ้า
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,   # แถบสลับสี
    showColumnStripes=False
)

ws.add_table(table)
wb.save(output)

print(f"✅ เสร็จสิ้น! ได้ข้อมูลทั้งหมด {len(df)} รายการ")
print(f"📁 ไฟล์: {output}")
